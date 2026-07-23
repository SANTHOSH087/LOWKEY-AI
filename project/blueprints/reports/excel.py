import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def render_report_xlsx(report: dict, period_label: str, report_type_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = report_type_label[:31]  # Excel sheet name limit

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="EFEBFF", end_color="EFEBFF", fill_type="solid")

    ws["A1"] = "Lowkey AI"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"{report_type_label} Report — {period_label}"
    ws["A2"].font = Font(italic=True, color="666666")

    # summary block
    row = 4
    ws.cell(row=row, column=1, value="Summary").font = bold
    row += 1
    for key, value in report["summary"].items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    headers = report["table_headers"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill
    header_row = row
    row += 1

    for data_row in report["table_rows"]:
        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if isinstance(value, float):
                # real ₹ number formatting on the cell, not just a string label —
                # this is what makes it usable in Excel formulas, unlike a text prefix
                cell.number_format = '"₹"#,##0.00'
        row += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
